from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
import pandas as pd
def pandas_type_to_sqlite_type(pandas_type):
    if "object" in pandas_type:
        return "TEXT"
    elif "int" in pandas_type:
        return "INTEGER"
    elif "float" in pandas_type:
        return "REAL"
    elif "datetime" in pandas_type:
        return "TIMESTAMP"
    else:
        return "TEXT"
    
def read_and_remove_strikethrough_content(excel_file_path: str, sheetname = None) -> pd.DataFrame:
    wb: Workbook = load_workbook(excel_file_path ,rich_text=True)
    ws = None
    if sheetname is None:
        ws = wb.active
    else:
        ws = wb.get_sheet_by_name(sheetname)
    data = []
    cols = []
    row_count = 0
    for row in ws.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            if(row_count == 0):
                cols.append(cell.value)
            cleaned_text = ""
            if(type(cell.value) == CellRichText):
                for text in cell.value:
                    if text.font.strike:
                        continue
                    else:
                        cleaned_text += text.text
            else:
                cleaned_text = cell.value
            row_data.append(cleaned_text)
        row_count += 1
        data.append(row_data)
    return pd.DataFrame(data, columns=cols)
